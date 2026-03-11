"""Router de Cadastros (Cursos, Projetos, Empresas) - MongoDB version"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional, List
from datetime import datetime, timezone
import uuid
import io
import re

from src.domain.entities import Usuario
from src.domain.value_objects import StatusPedido
from src.infrastructure.persistence.mongodb import db
from src.routers.dependencies import get_current_user

router = APIRouter(tags=["Cadastros"])

TIPOS_CURSO = [
    {"value": "tecnico", "label": "Escola Técnica"},
    {"value": "graduacao", "label": "Graduação"},
    {"value": "pos_graduacao", "label": "Pós-Graduação"},
    {"value": "qualificacao", "label": "Qualificação Profissional"},
    {"value": "aperfeicoamento", "label": "Aperfeiçoamento"},
    {"value": "livre", "label": "Curso Livre"},
]
MODALIDADES_CURSO = [
    {"value": "presencial", "label": "Presencial"},
    {"value": "ead", "label": "EAD"},
    {"value": "hibrido", "label": "Híbrido"},
]
AREAS_CURSO = [
    {"value": "tecnologia", "label": "Tecnologia da Informação"},
    {"value": "industria", "label": "Indústria e Manufatura"},
    {"value": "gestao", "label": "Gestão e Negócios"},
    {"value": "saude", "label": "Saúde"},
    {"value": "automacao", "label": "Automação e Mecatrônica"},
    {"value": "eletrica", "label": "Elétrica e Eletrônica"},
    {"value": "mecanica", "label": "Mecânica"},
    {"value": "quimica", "label": "Química"},
    {"value": "seguranca", "label": "Segurança do Trabalho"},
    {"value": "outros", "label": "Outros"},
]


class CursoRequest(BaseModel):
    nome: str = Field(..., min_length=3, max_length=200)
    descricao: Optional[str] = None
    tipo: Optional[str] = None
    modalidade: Optional[str] = None
    area: Optional[str] = None
    carga_horaria: Optional[str] = None
    duracao: Optional[str] = None

class ProjetoRequest(BaseModel):
    nome: str = Field(..., min_length=3, max_length=200)
    descricao: Optional[str] = None

class EmpresaRequest(BaseModel):
    nome: str = Field(..., min_length=3, max_length=200)
    cnpj: Optional[str] = None

class ImportacaoCursosResponse(BaseModel):
    total_linhas: int
    importados: int
    duplicados: int
    erros: int
    detalhes: List[dict]


def get_label(lista, value):
    if not value:
        return None
    for item in lista:
        if item["value"] == value:
            return item["label"]
    return value


def _curso_response(c):
    return {
        "id": c["id"], "nome": c["nome"], "descricao": c.get("descricao"),
        "tipo": c.get("tipo"), "tipo_label": get_label(TIPOS_CURSO, c.get("tipo")),
        "modalidade": c.get("modalidade"), "modalidade_label": get_label(MODALIDADES_CURSO, c.get("modalidade")),
        "area": c.get("area"), "area_label": get_label(AREAS_CURSO, c.get("area")),
        "carga_horaria": c.get("carga_horaria"), "duracao": c.get("duracao"),
        "ativo": c.get("ativo", True)
    }


# ==================== CURSOS ====================

@router.get("/cursos/opcoes", tags=["Cursos"])
async def listar_opcoes_cursos():
    return {"tipos": TIPOS_CURSO, "modalidades": MODALIDADES_CURSO, "areas": AREAS_CURSO}


@router.get("/cursos", tags=["Cursos"])
async def listar_cursos(
    ativo: Optional[bool] = None, tipo: Optional[str] = None,
    modalidade: Optional[str] = None, area: Optional[str] = None,
    busca: Optional[str] = None,
    current_user: Usuario = Depends(get_current_user)
):
    query = {}
    if ativo is not None:
        query["ativo"] = ativo
    if tipo:
        query["tipo"] = tipo
    if modalidade:
        query["modalidade"] = modalidade
    if area:
        query["area"] = area
    if busca:
        query["nome"] = {"$regex": busca, "$options": "i"}

    cursor = db.cursos.find(query, {"_id": 0}).sort([("tipo", 1), ("nome", 1)])
    cursos = await cursor.to_list(length=10000)
    return [_curso_response(c) for c in cursos]


@router.get("/cursos/estatisticas", tags=["Cursos"])
async def estatisticas_cursos(current_user: Usuario = Depends(get_current_user)):
    pipeline_tipo = [
        {"$match": {"ativo": {"$ne": False}}},
        {"$group": {"_id": {"$ifNull": ["$tipo", "sem_tipo"]}, "count": {"$sum": 1}}}
    ]
    por_tipo = {r["_id"]: r["count"] for r in await db.cursos.aggregate(pipeline_tipo).to_list(100)}

    pipeline_mod = [
        {"$match": {"ativo": {"$ne": False}}},
        {"$group": {"_id": {"$ifNull": ["$modalidade", "sem_modalidade"]}, "count": {"$sum": 1}}}
    ]
    por_modalidade = {r["_id"]: r["count"] for r in await db.cursos.aggregate(pipeline_mod).to_list(100)}

    total = await db.cursos.count_documents({"ativo": {"$ne": False}})
    return {"total": total, "por_tipo": por_tipo, "por_modalidade": por_modalidade}


@router.post("/cursos", status_code=201, tags=["Cursos"])
async def criar_curso(request: CursoRequest, current_user: Usuario = Depends(get_current_user)):
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    existing = await db.cursos.find_one({"nome": request.nome})
    if existing:
        raise HTTPException(409, "Curso já existe")

    doc = {"id": str(uuid.uuid4()), "nome": request.nome, "descricao": request.descricao,
           "tipo": request.tipo, "modalidade": request.modalidade, "area": request.area,
           "carga_horaria": request.carga_horaria, "duracao": request.duracao,
           "ativo": True, "created_at": datetime.now(timezone.utc).isoformat(),
           "updated_at": datetime.now(timezone.utc).isoformat()}
    await db.cursos.insert_one(doc)
    return _curso_response(doc)


@router.put("/cursos/{curso_id}", tags=["Cursos"])
async def atualizar_curso(curso_id: str, request: CursoRequest, current_user: Usuario = Depends(get_current_user)):
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    result = await db.cursos.update_one({"id": curso_id}, {"$set": {
        "nome": request.nome, "descricao": request.descricao, "tipo": request.tipo,
        "modalidade": request.modalidade, "area": request.area,
        "carga_horaria": request.carga_horaria, "duracao": request.duracao,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }})
    if result.matched_count == 0:
        raise HTTPException(404, "Curso não encontrado")
    doc = await db.cursos.find_one({"id": curso_id}, {"_id": 0})
    return _curso_response(doc)


@router.patch("/cursos/{curso_id}/ativar", tags=["Cursos"])
async def ativar_curso(curso_id: str, current_user: Usuario = Depends(get_current_user)):
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    result = await db.cursos.update_one({"id": curso_id}, {"$set": {"ativo": True}})
    if result.matched_count == 0:
        raise HTTPException(404, "Curso não encontrado")
    return {"message": "Curso ativado com sucesso"}


@router.delete("/cursos/{curso_id}", tags=["Cursos"])
async def deletar_curso(curso_id: str, current_user: Usuario = Depends(get_current_user)):
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    result = await db.cursos.update_one({"id": curso_id}, {"$set": {"ativo": False}})
    if result.matched_count == 0:
        raise HTTPException(404, "Curso não encontrado")
    return {"message": "Curso desativado com sucesso"}


# ==================== PROJETOS ====================

@router.get("/projetos", tags=["Projetos"])
async def listar_projetos(ativo: Optional[bool] = None, current_user: Usuario = Depends(get_current_user)):
    query = {}
    if ativo is not None:
        query["ativo"] = ativo
    projetos = await db.projetos.find(query, {"_id": 0}).sort("nome", 1).to_list(1000)
    return [{"id": p["id"], "nome": p["nome"], "descricao": p.get("descricao"), "ativo": p.get("ativo", True)} for p in projetos]


@router.post("/projetos", status_code=201, tags=["Projetos"])
async def criar_projeto(request: ProjetoRequest, current_user: Usuario = Depends(get_current_user)):
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    if await db.projetos.find_one({"nome": request.nome}):
        raise HTTPException(409, "Projeto já existe")
    doc = {"id": str(uuid.uuid4()), "nome": request.nome, "descricao": request.descricao, "ativo": True,
           "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()}
    await db.projetos.insert_one(doc)
    return {"id": doc["id"], "nome": doc["nome"], "descricao": doc.get("descricao"), "ativo": True}


@router.put("/projetos/{projeto_id}", tags=["Projetos"])
async def atualizar_projeto(projeto_id: str, request: ProjetoRequest, current_user: Usuario = Depends(get_current_user)):
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    result = await db.projetos.update_one({"id": projeto_id}, {"$set": {"nome": request.nome, "descricao": request.descricao, "updated_at": datetime.now(timezone.utc).isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(404, "Projeto não encontrado")
    doc = await db.projetos.find_one({"id": projeto_id}, {"_id": 0})
    return {"id": doc["id"], "nome": doc["nome"], "descricao": doc.get("descricao"), "ativo": doc.get("ativo", True)}


@router.delete("/projetos/{projeto_id}", tags=["Projetos"])
async def deletar_projeto(projeto_id: str, current_user: Usuario = Depends(get_current_user)):
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    result = await db.projetos.update_one({"id": projeto_id}, {"$set": {"ativo": False}})
    if result.matched_count == 0:
        raise HTTPException(404, "Projeto não encontrado")
    return {"message": "Projeto desativado com sucesso"}


# ==================== EMPRESAS ====================

@router.get("/empresas", tags=["Empresas"])
async def listar_empresas(ativo: Optional[bool] = None, current_user: Usuario = Depends(get_current_user)):
    query = {}
    if ativo is not None:
        query["ativo"] = ativo
    empresas = await db.empresas.find(query, {"_id": 0}).sort("nome", 1).to_list(1000)
    return [{"id": e["id"], "nome": e["nome"], "cnpj": e.get("cnpj"), "ativo": e.get("ativo", True)} for e in empresas]


@router.post("/empresas", status_code=201, tags=["Empresas"])
async def criar_empresa(request: EmpresaRequest, current_user: Usuario = Depends(get_current_user)):
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    if await db.empresas.find_one({"nome": request.nome}):
        raise HTTPException(409, "Empresa já existe")
    doc = {"id": str(uuid.uuid4()), "nome": request.nome, "cnpj": request.cnpj, "ativo": True,
           "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()}
    await db.empresas.insert_one(doc)
    return {"id": doc["id"], "nome": doc["nome"], "cnpj": doc.get("cnpj"), "ativo": True}


@router.put("/empresas/{empresa_id}", tags=["Empresas"])
async def atualizar_empresa(empresa_id: str, request: EmpresaRequest, current_user: Usuario = Depends(get_current_user)):
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    result = await db.empresas.update_one({"id": empresa_id}, {"$set": {"nome": request.nome, "cnpj": request.cnpj, "updated_at": datetime.now(timezone.utc).isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(404, "Empresa não encontrada")
    doc = await db.empresas.find_one({"id": empresa_id}, {"_id": 0})
    return {"id": doc["id"], "nome": doc["nome"], "cnpj": doc.get("cnpj"), "ativo": doc.get("ativo", True)}


@router.delete("/empresas/{empresa_id}", tags=["Empresas"])
async def deletar_empresa(empresa_id: str, current_user: Usuario = Depends(get_current_user)):
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    result = await db.empresas.update_one({"id": empresa_id}, {"$set": {"ativo": False}})
    if result.matched_count == 0:
        raise HTTPException(404, "Empresa não encontrada")
    return {"message": "Empresa desativada com sucesso"}


# ==================== IMPORTAÇÃO CURSOS ====================

@router.get("/cursos/importacao/template", tags=["Cursos - Importação"])
async def download_template_cursos():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cursos"
    headers = ["Código", "Nome", "Descrição", "Tipo", "Modalidade", "Área", "Carga Horária", "Duração"]
    header_fill = PatternFill(start_color="004587", end_color="004587", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=template_cursos_senai.xlsx"})


@router.post("/cursos/importacao/validar", tags=["Cursos - Importação"])
async def validar_importacao_cursos(arquivo: UploadFile = File(...), current_user: Usuario = Depends(get_current_user)):
    import pandas as pd
    if not arquivo.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Arquivo deve ser Excel")
    conteudo = await arquivo.read()
    df = pd.read_excel(io.BytesIO(conteudo))
    colunas_lower = {col.lower().strip(): col for col in df.columns}
    col_nome = colunas_lower.get('nome')
    if not col_nome:
        raise HTTPException(400, "Coluna 'Nome' é obrigatória")

    existing = await db.cursos.distinct("nome")
    cursos_existentes = {c.lower() for c in existing}

    validados, duplicados, erros = [], [], []
    for idx, row in df.iterrows():
        nome = str(row[col_nome]).strip() if pd.notna(row[col_nome]) else None
        if not nome or nome.lower() == 'nan':
            erros.append({"linha": idx + 2, "erro": "Nome vazio"})
            continue
        if nome.lower() in cursos_existentes:
            duplicados.append({"linha": idx + 2, "nome": nome, "motivo": "Já existe"})
            continue
        validados.append({"linha": idx + 2, "nome": nome})

    return {"total_linhas": len(df), "validos": len(validados), "duplicados": len(duplicados),
            "erros": len(erros), "preview": validados[:10], "duplicados_lista": duplicados[:10], "erros_lista": erros[:10]}


@router.post("/cursos/importacao/executar", tags=["Cursos - Importação"])
async def executar_importacao_cursos(arquivo: UploadFile = File(...), current_user: Usuario = Depends(get_current_user)):
    import pandas as pd
    if current_user.role.value not in ("admin",):
        raise HTTPException(403, "Sem permissão")
    if not arquivo.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Arquivo deve ser Excel")
    conteudo = await arquivo.read()
    df = pd.read_excel(io.BytesIO(conteudo))
    colunas_lower = {col.lower().strip(): col for col in df.columns}
    col_nome = colunas_lower.get('nome')
    if not col_nome:
        raise HTTPException(400, "Coluna 'Nome' é obrigatória")

    existing = await db.cursos.distinct("nome")
    cursos_existentes = {c.lower() for c in existing}

    importados, duplicados, erros, detalhes = 0, 0, 0, []
    nomes_importados = set()
    batch = []

    for idx, row in df.iterrows():
        nome = str(row[col_nome]).strip() if pd.notna(row[col_nome]) else None
        if not nome or nome.lower() == 'nan':
            erros += 1
            continue
        if nome.lower() in cursos_existentes or nome.lower() in nomes_importados:
            duplicados += 1
            continue
        batch.append({"id": str(uuid.uuid4()), "nome": nome, "ativo": True,
                       "created_at": datetime.now(timezone.utc).isoformat()})
        nomes_importados.add(nome.lower())
        importados += 1

    if batch:
        await db.cursos.insert_many(batch)

    return {"total_linhas": len(df), "importados": importados, "duplicados": duplicados, "erros": erros, "detalhes": detalhes[:100]}


# ==================== DADOS AUXILIARES ====================

@router.get("/status-pedido", tags=["Auxiliares"])
async def listar_status():
    return [{"value": s.value, "label": s.label} for s in StatusPedido]
