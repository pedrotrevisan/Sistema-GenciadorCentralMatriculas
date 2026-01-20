"""Exportador XLSX para TOTVS - Padrão Strategy"""
from abc import ABC, abstractmethod
from typing import List, BinaryIO
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import csv

from src.domain.entities import PedidoMatricula


class IExportadorArquivo(ABC):
    """Interface para exportadores de arquivo - Strategy Pattern"""

    @abstractmethod
    def exportar(self, pedidos: List[PedidoMatricula]) -> BytesIO:
        """Exporta pedidos para o formato específico"""
        pass

    @abstractmethod
    def get_content_type(self) -> str:
        """Retorna o content-type do arquivo"""
        pass

    @abstractmethod
    def get_extension(self) -> str:
        """Retorna a extensão do arquivo"""
        pass


class ExportadorXLSXTOTVS(IExportadorArquivo):
    """Exportador para Excel no formato TOTVS - ORDEM EXATA DO TOTVS"""

    def exportar(self, pedidos: List[PedidoMatricula]) -> BytesIO:
        """Exporta pedidos para XLSX no layout TOTVS"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Matrículas TOTVS"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="004587", end_color="004587", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Cabeçalhos na ORDEM EXATA do TOTVS
        headers = [
            "CPF",                    # 1
            "RG",                     # 2
            "Emissão RG",            # 3
            "Orgão emissor",         # 4
            "Nome completo do aluno", # 5
            "Estado Natal/UF",       # 6
            "Naturalidade",          # 7
            "Data de nascimento",    # 8
            "Sexo",                  # 9
            "Cor/Raça",              # 10
            "Grau de Instrução",     # 11
            "Endereço",              # 12
            "Nº",                    # 13
            "Complemento",           # 14
            "Bairro",                # 15
            "Cidade",                # 16
            "Telefone",              # 17
            "E-mail",                # 18
            "Nome do pai",           # 19
            "Nome da mãe",           # 20
            # Campos extras do sistema
            "Protocolo",             # 21
            "Curso",                 # 22
            "Projeto/Empresa",       # 23
            "Consultor",             # 24
        ]

        # Aplica cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Dados
        row = 2
        
        for pedido in pedidos:
            for aluno in pedido.alunos:
                # Dados na ORDEM EXATA do TOTVS
                ws.cell(row=row, column=1, value=aluno.cpf.formatado()).border = thin_border
                ws.cell(row=row, column=2, value=aluno.rg).border = thin_border
                ws.cell(row=row, column=3, value=aluno.rg_data_emissao or "").border = thin_border
                ws.cell(row=row, column=4, value=aluno.rg_orgao_emissor).border = thin_border
                ws.cell(row=row, column=5, value=aluno.nome).border = thin_border
                ws.cell(row=row, column=6, value=aluno.naturalidade_uf or "").border = thin_border
                ws.cell(row=row, column=7, value=aluno.naturalidade or "").border = thin_border
                ws.cell(row=row, column=8, value=aluno.data_nascimento.strftime("%d/%m/%Y")).border = thin_border
                ws.cell(row=row, column=9, value=aluno.sexo or "").border = thin_border
                ws.cell(row=row, column=10, value=aluno.cor_raca or "").border = thin_border
                ws.cell(row=row, column=11, value=aluno.grau_instrucao or "").border = thin_border
                ws.cell(row=row, column=12, value=aluno.endereco_logradouro).border = thin_border
                ws.cell(row=row, column=13, value=aluno.endereco_numero).border = thin_border
                ws.cell(row=row, column=14, value=aluno.endereco_complemento or "").border = thin_border
                ws.cell(row=row, column=15, value=aluno.endereco_bairro).border = thin_border
                ws.cell(row=row, column=16, value=aluno.endereco_cidade).border = thin_border
                ws.cell(row=row, column=17, value=aluno.telefone.formatado()).border = thin_border
                ws.cell(row=row, column=18, value=aluno.email.valor).border = thin_border
                ws.cell(row=row, column=19, value=aluno.nome_pai or "").border = thin_border
                ws.cell(row=row, column=20, value=aluno.nome_mae or "").border = thin_border
                # Campos extras
                ws.cell(row=row, column=21, value=pedido.numero_protocolo or pedido.id[:8]).border = thin_border
                ws.cell(row=row, column=22, value=pedido.curso_nome).border = thin_border
                ws.cell(row=row, column=23, value=pedido.projeto_nome or pedido.empresa_nome or "").border = thin_border
                ws.cell(row=row, column=24, value=pedido.consultor_nome).border = thin_border
                row += 1

        # Ajusta largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Salva em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_extension(self) -> str:
        return "xlsx"


class ExportadorCSVTOTVS(IExportadorArquivo):
    """Exportador para CSV no formato TOTVS - ORDEM EXATA DO TOTVS"""

    def exportar(self, pedidos: List[PedidoMatricula]) -> BytesIO:
        """Exporta pedidos para CSV no layout TOTVS"""
        output = BytesIO()
        
        # Cabeçalhos na ORDEM EXATA do TOTVS
        headers = [
            "CPF", "RG", "Emissão RG", "Orgão emissor", "Nome completo do aluno",
            "Estado Natal/UF", "Naturalidade", "Data de nascimento", "Sexo",
            "Cor/Raça", "Grau de Instrução", "Endereço", "Nº", "Complemento",
            "Bairro", "Cidade", "Telefone", "E-mail", "Nome do pai", "Nome da mãe",
            "Protocolo", "Curso", "Projeto/Empresa", "Consultor"
        ]
        
        # Escreve em modo texto
        import io
        text_output = io.StringIO()
        writer = csv.writer(text_output, delimiter=';', quoting=csv.QUOTE_ALL)
        writer.writerow(headers)
        
        for pedido in pedidos:
            for aluno in pedido.alunos:
                writer.writerow([
                    aluno.cpf.formatado(),
                    aluno.rg,
                    aluno.rg_data_emissao or "",
                    aluno.rg_orgao_emissor,
                    aluno.nome,
                    aluno.naturalidade_uf or "",
                    aluno.naturalidade or "",
                    aluno.data_nascimento.strftime("%d/%m/%Y"),
                    aluno.sexo or "",
                    aluno.cor_raca or "",
                    aluno.grau_instrucao or "",
                    aluno.endereco_logradouro,
                    aluno.endereco_numero,
                    aluno.endereco_complemento or "",
                    aluno.endereco_bairro,
                    aluno.endereco_cidade,
                    aluno.telefone.formatado(),
                    aluno.email.valor,
                    aluno.nome_pai or "",
                    aluno.nome_mae or "",
                    pedido.numero_protocolo or pedido.id[:8],
                    pedido.curso_nome,
                    pedido.projeto_nome or pedido.empresa_nome or "",
                    pedido.consultor_nome
                ])
        
        # Converte para bytes
        output.write(text_output.getvalue().encode('utf-8-sig'))
        output.seek(0)
        return output

    def get_content_type(self) -> str:
        return "text/csv"

    def get_extension(self) -> str:
        return "csv"


class ExportadorFactory:
    """Factory para criar exportadores - Factory Pattern"""

    _exportadores = {
        "xlsx": ExportadorXLSXTOTVS,
        "csv": ExportadorCSVTOTVS
    }

    @classmethod
    def criar(cls, formato: str) -> IExportadorArquivo:
        """Cria exportador baseado no formato"""
        formato = formato.lower()
        if formato not in cls._exportadores:
            raise ValueError(f"Formato não suportado: {formato}")
        return cls._exportadores[formato]()

    @classmethod
    def formatos_disponiveis(cls) -> list:
        """Retorna lista de formatos disponíveis"""
        return list(cls._exportadores.keys())
